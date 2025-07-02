import re
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
import csv
from .serializers import ImportSerializer, SubSubCategorySmallSerializer
from .models import Product, ProductCategory, ProductSubCategory, Size, ProductImage
import openpyxl
from rest_framework.parsers import MultiPartParser
from django.core.files.base import ContentFile
import requests
import os
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from django.utils.text import slugify
from django.core.files.temp import NamedTemporaryFile
from django.core.files import File
import pandas as pd
import io
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl import Workbook
from django.http import HttpResponse
from rest_framework import generics
from .models import Product, ProductCategory, ProductSubCategory, Wishlist, ProductReview, ProductImage, Size, ProductSubSubCategory
from .serializers import ProductSerializer, CategorySerializer, SubCategorySerializer, ProductDetailSerializer, WishlistSerializer, ProductReviewDetailSerializer, ProductReviewSerializer, ProductImageSerializer, ProductImageSmallSerializer, SizeSerializer, ImportSerializer, ProductListSerializer, CategorySmallSerializer, SubCategorySmallSerializer, SubSubCategorySerializer, SubSubCategoryListSerializer
from django_filters import rest_framework as django_filters
from rest_framework import filters
from django.http import Http404
from rest_framework.response import Response
from rest_framework import status
from rest_framework.views import APIView
from rest_framework.pagination import PageNumberPagination
# Create your views here.


class CustomPagination(PageNumberPagination):
    page_size = 18
    page_size_query_param = 'page_size'
    max_page_size = 100


class CategoryListCreateView(generics.ListCreateAPIView):
    queryset = ProductCategory.objects.only(
        'id', 'name', 'slug', 'description', 'image')
    serializer_class = CategorySerializer

    def get_serializer_class(self):
        if self.request.method == 'GET':
            return CategorySmallSerializer
        return CategorySerializer


class CategoryDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = ProductCategory.objects.only(
        'id', 'name', 'slug', 'description', 'image')
    serializer_class = CategorySerializer
    lookup_field = 'slug'

    def get_serializer_class(self):
        if self.request.method == 'GET':
            return CategorySmallSerializer
        return CategorySerializer


class SubCategoryFilter(django_filters.FilterSet):
    category = django_filters.CharFilter(
        field_name='category__name', lookup_expr='icontains')

    class Meta:
        model = ProductSubCategory
        fields = ['category']


class SubCategoryListCreateView(generics.ListCreateAPIView):
    queryset = ProductSubCategory.objects.only(
        'id', 'name', 'slug', 'category', 'image', 'description').select_related('category')
    serializer_class = SubCategorySerializer
    filter_backends = [django_filters.DjangoFilterBackend]
    filterset_class = SubCategoryFilter

    def get_serializer_class(self):
        if self.request.method == 'GET':
            return SubCategorySmallSerializer
        return SubCategorySerializer

    def get_queryset(self):
        category_slug = self.request.query_params.get('category_slug')
        if category_slug:
            return ProductSubCategory.objects.filter(category__slug=category_slug).only('id', 'name', 'slug', 'category', 'image').select_related('category')
        return ProductSubCategory.objects.only(
            'id', 'name', 'slug', 'category', 'image').select_related('category')


class SubCategoryRetrieveUpdateDestroyView(generics.RetrieveUpdateDestroyAPIView):
    queryset = ProductSubCategory.objects.only(
        'id', 'name', 'slug', 'category', 'image').select_related('category')
    serializer_class = SubCategorySerializer
    lookup_field = 'slug'

    def get_serializer_class(self):
        if self.request.method == 'GET':
            return SubCategorySmallSerializer
        return SubCategorySerializer


class SubSubCategoryListCreateView(generics.ListCreateAPIView):
    queryset = ProductSubSubCategory.objects.only(
        'id', 'name', 'slug', 'subcategory', 'image', 'description').select_related('subcategory')
    serializer_class = SubSubCategorySerializer

    def get_serializer_class(self):
        if self.request.method == 'GET':
            return SubSubCategorySmallSerializer
        return SubSubCategorySerializer

    def get_queryset(self):
        subcategory_slug = self.request.query_params.get('subcategory_slug')
        if subcategory_slug:
            return ProductSubSubCategory.objects.filter(subcategory__slug=subcategory_slug).only('id', 'name', 'slug', 'subcategory', 'image').select_related('subcategory')
        return ProductSubSubCategory.objects.only(
            'id', 'name', 'slug', 'subcategory', 'image').select_related('subcategory')


class SubSubCategoryRetrieveUpdateDestroyView(generics.RetrieveUpdateDestroyAPIView):
    queryset = ProductSubSubCategory.objects.only(
        'id', 'name', 'slug', 'subcategory', 'image').select_related('subcategory')
    serializer_class = SubSubCategorySerializer
    lookup_field = 'slug'


class SizeListCreateView(generics.ListCreateAPIView):
    queryset = Size.objects.only('id', 'name', 'description')
    serializer_class = SizeSerializer


class SizeRetrieveUpdateDestroyView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Size.objects.only('id', 'name', 'description')
    serializer_class = SizeSerializer
    lookup_field = 'id'


class ProductImageListCreateView(generics.ListCreateAPIView):
    queryset = ProductImage.objects.all()
    serializer_class = ProductImageSmallSerializer


class ProductImageRetrieveUpdateDestroyView(generics.RetrieveUpdateDestroyAPIView):
    queryset = ProductImage.objects.all()
    lookup_field = 'id'

    def get_object(self):
        try:
            return ProductImage.objects.get(id=self.kwargs['id'])
        except ProductImage.DoesNotExist:
            raise Http404("Product image not found")

    def get_serializer_class(self):
        if self.request.method == 'GET':
            return ProductImageSmallSerializer
        return ProductImageSerializer


class ProductFilter(django_filters.FilterSet):
    name = django_filters.CharFilter(lookup_expr='icontains')
    min_price = django_filters.NumberFilter(
        field_name='price', lookup_expr='gte')
    max_price = django_filters.NumberFilter(
        field_name='price', lookup_expr='lte')
    category = django_filters.CharFilter(
        field_name='category__slug', lookup_expr='exact')
    subcategory = django_filters.CharFilter(
        field_name='subcategory__slug', lookup_expr='exact')
    subsubcategory = django_filters.CharFilter(
        field_name='subsubcategory__slug', lookup_expr='exact')
    is_popular = django_filters.BooleanFilter(
        field_name='is_popular', lookup_expr='exact')
    is_featured = django_filters.BooleanFilter(
        field_name='is_featured', lookup_expr='exact')

    class Meta:
        model = Product
        fields = ['name', 'min_price', 'max_price', 'category',
                  'subcategory', 'subsubcategory', 'is_popular', 'is_featured']


class ProductListCreateView(generics.ListCreateAPIView):
    queryset = Product.objects.only(
        'id', 'name', 'slug', 'market_price', 'price', 'is_popular', 'is_featured', 'stock',
        'thumbnail_image', 'thumbnail_image_alt_description', 'category', 'subcategory', 'subsubcategory', 'is_featured', 'is_popular', 'is_active'
    ).select_related('category', 'subcategory', 'subsubcategory').order_by('-created_at')
    serializer_class = ProductSerializer
    filter_backends = [django_filters.DjangoFilterBackend,
                       filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['name']
    ordering_fields = ['name', 'price', '-name', '-price', '-created_at']
    filterset_class = ProductFilter
    pagination_class = CustomPagination
    parser_classes = [MultiPartParser, FormParser, JSONParser]

    def get_serializer_class(self):
        if self.request.method == 'GET':
            return ProductListSerializer
        return ProductSerializer


class ProductDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Product.objects.all()
    lookup_field = 'slug'

    def get_object(self):
        subsubcategory_slug = self.kwargs.get('subsubcategory_slug')
        slug = self.kwargs.get('slug')

        try:
            return Product.objects.get(subsubcategory__slug=subsubcategory_slug, slug=slug)
        except Product.DoesNotExist:
            raise Http404("Product not found")

    def get_serializer_class(self):
        if self.request.method == 'GET':
            return ProductDetailSerializer
        return ProductSerializer


class SimilarProductsView(generics.ListAPIView):
    serializer_class = ProductListSerializer

    def get_queryset(self):
        slug = self.kwargs.get('slug')
        try:
            product = Product.objects.only('id', 'category').get(slug=slug)
            # Get products from the same category, excluding the current product
            similar_products = Product.objects.only(
                'id', 'name', 'slug', 'market_price', 'price', 'is_popular', 'is_featured',
                'thumbnail_image', 'thumbnail_image_alt_description'
            ).filter(
                category=product.category,
                is_active=True
            ).exclude(
                id=product.id
            ).order_by('-created_at')[:3]
            return similar_products
        except Product.DoesNotExist:
            return Product.objects.none()


class WishlistListCreateView(generics.ListCreateAPIView):
    queryset = Wishlist.objects.only(
        'id', 'user', 'product', 'created_at', 'updated_at'
    ).select_related(
        'user',
        'product'
    )
    serializer_class = WishlistSerializer

    def get_queryset(self):
        return Wishlist.objects.filter(user=self.request.user)

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)


class WishlistRetrieveUpdateDestroyView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Wishlist.objects.only(
        'id', 'user', 'product', 'created_at', 'updated_at'
    ).select_related(
        'user',
        'product'
    )
    serializer_class = WishlistSerializer

    def get_object(self):
        try:
            return Wishlist.objects.get(user=self.request.user)
        except Wishlist.DoesNotExist:
            raise Http404("Wishlist not found")

    def delete(self, request, *args, **kwargs):
        try:
            id = self.kwargs.get('id')
            wishlist = Wishlist.objects.get(
                user=self.request.user, id=id)
            wishlist.delete()
            return Response({"message": "Product removed from wishlist"}, status=status.HTTP_204_NO_CONTENT)
        except Wishlist.DoesNotExist:
            raise Http404("Wishlist not found")


class ProductReviewFilter(django_filters.FilterSet):
    rating = django_filters.NumberFilter(
        field_name='rating', lookup_expr='exact')

    class Meta:
        model = ProductReview
        fields = ['rating']


class ProductReviewView(generics.ListCreateAPIView):
    queryset = ProductReview.objects.only(
        'id', 'product', 'user', 'review', 'rating', 'created_at'
    ).select_related('product', 'user').order_by('-created_at')
    serializer_class = ProductReviewSerializer
    pagination_class = CustomPagination
    filter_backends = [django_filters.DjangoFilterBackend,]
    filterset_class = ProductReviewFilter

    def get_serializer_class(self):
        if self.request.method == 'GET':
            return ProductReviewDetailSerializer
        return ProductReviewSerializer

    def get_queryset(self):
        slug = self.request.query_params.get('slug')
        try:
            product = Product.objects.only('id').get(slug=slug)
            return ProductReview.objects.filter(product=product)
        except Product.DoesNotExist:
            return ProductReview.objects.all()

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)


class ProductReviewRetrieveUpdateDestroyView(generics.RetrieveUpdateDestroyAPIView):
    queryset = ProductReview.objects.only(
        'id', 'product', 'user', 'review', 'rating', 'created_at', 'updated_at'
    ).select_related('product', 'user')
    serializer_class = ProductReviewSerializer
    lookup_field = 'id'


class ProductExcelExportWithDropdownAPIView(APIView):
    def get(self, request, format=None):
        categories = list(
            ProductCategory.objects.values_list('name', flat=True))

        # Get subcategories with their category names
        sub_categories_with_category = []
        for subcategory in ProductSubCategory.objects.select_related('category').all():
            sub_categories_with_category.append(
                f"{subcategory.name} ({subcategory.category.name})")

        # Get sub-subcategories with their subcategory names
        sub_sub_categories_with_subcategory = []
        for subsubcategory in ProductSubSubCategory.objects.select_related('subcategory').all():
            sub_sub_categories_with_subcategory.append(
                f"{subsubcategory.name} ({subsubcategory.subcategory.name})")

        sizes = list(Size.objects.values_list('name', flat=True))

        output = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Products"

        # Add a reference sheet for dropdowns
        ref_ws = wb.create_sheet(title="DropdownLists")
        for idx, value in enumerate(categories, 1):
            ref_ws.cell(row=idx, column=1, value=value)
        for idx, value in enumerate(sub_categories_with_category, 1):
            ref_ws.cell(row=idx, column=2, value=value)
        for idx, value in enumerate(sub_sub_categories_with_subcategory, 1):
            ref_ws.cell(row=idx, column=3, value=value)

        # Define headers
        headers = [
            'Product Name', 'Category', 'Sub Category', 'Sub Sub Category', 'Price', 'Market Price', 'Discount',
            'Product Stock', 'Is popular', 'Is featured', 'Description', 'Highlight Description', 'Extra Description', 'Specifications', 'Meta Title', 'Meta Description',
            'Size', 'Thumbnail image', 'Thumbnail Image Alt Description',
            'Color', 'Stock (Color)', 'Images', 'Image Alt Description'
        ]
        ws.append(headers)

        # Sample common data (only on first row)
        common_data = {
            'Product Name': 'Product Name',
            'Category': 'Category',
            'Sub Category': 'Sub Category (Category)',
            'Sub Sub Category': 'Sub Sub Category (Sub Category)',
            'Price': 100,
            'Market Price': 100,
            'Discount': 100,
            'Product Stock': 100,
            'Is popular': 'TRUE',
            'Is featured': 'FALSE',
            'Description': 'Product Description',
            'Highlight Description': 'Product Highlight Description',
            'Extra Description': 'Product Extra Description',
            'Specifications': 'Product Specifications',
            'Meta Title': 'Product Meta Title',
            'Meta Description': 'Product Meta Description',
            'Size': 'S, M, L',
            'Thumbnail image': 'https://example.com/thumbnail.jpg',
            'Thumbnail Image Alt Description': 'Thumbnail Alt Description',
        }

        # Sample image rows with 3 variants (Color 1, 2, 3)
        for i in range(3):
            row = []
            for h in headers:
                if h in ['Color', 'Stock (Color)', 'Images', 'Image Alt Description']:
                    continue
                # Fill only on first row
                row.append(common_data[h] if i == 0 else '')

            row.append(f'Color {i+1}')
            row.append(10 * (i+1))  # Stock
            row.append(f'https://example.com/image{i+1}.jpg')
            row.append(f'Image Alt Description {i+1}')

            ws.append(row)

        # Data validation using reference sheet
        from openpyxl.utils import quote_sheetname

        if categories:
            dv = DataValidation(
                type="list",
                formula1=f"={quote_sheetname('DropdownLists')}!$A$1:$A${len(categories)}",
                allow_blank=True
            )
            dv.add("B2:B100")
            ws.add_data_validation(dv)

        if sub_categories_with_category:
            dv = DataValidation(
                type="list",
                formula1=f"={quote_sheetname('DropdownLists')}!$B$1:$B${len(sub_categories_with_category)}",
                allow_blank=True
            )
            dv.add("C2:C100")
            ws.add_data_validation(dv)

        if sub_sub_categories_with_subcategory:
            dv = DataValidation(
                type="list",
                formula1=f"={quote_sheetname('DropdownLists')}!$C$1:$C${len(sub_sub_categories_with_subcategory)}",
                allow_blank=True
            )
            dv.add("D2:D100")
            ws.add_data_validation(dv)

        # Dropdown for booleans
        for col in ['H', 'I']:
            dv = DataValidation(
                type="list", formula1='"TRUE,FALSE"', allow_blank=True)
            dv.add(f"{col}2:{col}100")
            ws.add_data_validation(dv)

        # Auto column width
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_len + 2

        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="products_export.xlsx"'
        return response


class CategoryExcelUploadView(APIView):
    serializer_class = ImportSerializer
    parser_classes = [MultiPartParser]

    def post(self, request, *args, **kwargs):
        excel_file = request.FILES.get('file')
        if not excel_file:
            return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            df = pd.read_excel(excel_file)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

        for _, row in df.iterrows():
            cat1_name = str(row.get('Category①')).strip()
            cat2_name = str(row.get('Category②')).strip()
            cat3_name = str(row.get('Category③')).strip(
            ) if pd.notna(row.get('Category③')) else None

            # Create or get Category①
            cat1, _ = ProductCategory.objects.get_or_create(name=cat1_name)

            # Create or get Category②
            cat2, _ = ProductSubCategory.objects.get_or_create(
                name=cat2_name,
                category=cat1
            )

            # Create or get Category③ if present
            if cat3_name:
                ProductSubSubCategory.objects.get_or_create(
                    name=cat3_name,
                    subcategory=cat2
                )

        return Response({'message': 'Categories imported successfully'}, status=status.HTTP_201_CREATED)


def clean_name(name):
    """Removes content within parentheses and trims spaces."""
    return re.sub(r'\s*\(.*?\)', '', name or '').strip()


def get_or_create_category_hierarchy(row):
    category_name = clean_name(row.get('Category'))
    subcategory_name = clean_name(row.get('Sub Category'))
    subsubcategory_name = clean_name(row.get('Sub Sub Category'))

    category, _ = ProductCategory.objects.get_or_create(name=category_name)
    subcategory, _ = ProductSubCategory.objects.get_or_create(
        name=subcategory_name, category=category)
    subsubcategory, _ = ProductSubSubCategory.objects.get_or_create(
        name=subsubcategory_name, subcategory=subcategory)
    return category, subcategory, subsubcategory


def assign_sizes(product, size_str):
    sizes = [s.strip() for s in str(size_str or '').split(',') if s.strip()]
    for size_name in sizes:
        size_obj, _ = Size.objects.get_or_create(name=size_name)
        product.size.add(size_obj)


def save_image_from_url(instance, field_name, url, fallback_name):
    if url:
        try:
            response = requests.get(url)
            if response.status_code == 200:
                filename = url.split("/")[-1].split("?")[0] or fallback_name
                getattr(instance, field_name).save(
                    filename, ContentFile(response.content), save=True)
        except Exception as e:
            print(
                f"[{field_name}] Failed for {getattr(instance, 'name', 'unknown')}: {e}")


def create_product_from_row(row, category, subcategory, subsubcategory):
    product = Product.objects.create(
        name=row.get('Product Name'),
        slug=slugify(row.get('Product Name')),
        price=row.get('Price') or 0,
        market_price=row.get('Market Price') or 0,
        discount=row.get('Discount') or 0,
        stock=row.get('Product Stock') or 0,
        is_popular=str(row.get('Is popular')).lower() == 'true',
        is_featured=str(row.get('Is featured')).lower() == 'true',
        description=row.get('Description') or '',
        highlight_description=row.get('Highlight Description') or row.get(
            'Product Hightlight Description') or '',
        extra_description=row.get('Extra Description') or row.get(
            'Product Extra Description') or '',
        specifications=row.get('Specifications') or row.get(
            'Product Specifications') or '',
        meta_title=row.get('Meta Title') or row.get(
            'Product Meta Title') or '',
        meta_description=row.get('Meta Description') or row.get(
            'Product Meta Description') or '',
        thumbnail_image_alt_description=row.get(
            'Thumbnail Image Alt Description') or '',
        category=category,
        subcategory=subcategory,
        subsubcategory=subsubcategory
    )
    assign_sizes(product, row.get('Size'))
    save_image_from_url(product, 'thumbnail_image', row.get(
        'Thumbnail image'), f"{slugify(product.name)}-thumb.jpg")
    return product


def create_product_image(product, img_data):
    img_url = img_data.get('Images')
    if img_url:
        try:
            response = requests.get(img_url)
            if response.status_code == 200:
                image_file = ContentFile(response.content)
                image_filename = img_url.split(
                    "/")[-1].split("?")[0] or f"{slugify(product.name)}.jpg"
                image_instance = ProductImage.objects.create(
                    product=product,
                    color=img_data.get('Color', ''),
                    stock=img_data.get('Stock', 0) or img_data.get(
                        'Stock (Color)', 0) or 0,
                    image_alt_description=img_data.get(
                        'Image Alt Description', '') or img_data.get('Image Alt Description') or '',
                )
                image_instance.image.save(
                    image_filename, image_file, save=True)
        except Exception as e:
            print(f"[Image] Failed for {product.name}: {e}")


class ProductExcelImportAPIView(APIView):
    serializer_class = ImportSerializer
    parser_classes = [MultiPartParser]

    def post(self, request, format=None):
        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'No file uploaded'}, status=400)

        file_name = file.name.lower()
        try:
            if file_name.endswith('.csv'):
                rows = self.read_csv(file)
            elif file_name.endswith(('.xlsx', '.xls')):
                rows = self.read_excel(file)
            else:
                return Response({'error': 'Unsupported file type. Use CSV or Excel.'}, status=400)
            return self.process_rows(rows)
        except Exception as e:
            return Response({'error': str(e)}, status=500)

    def read_csv(self, file):
        decoded_file = file.read().decode('utf-8-sig')  # ✅ decode with BOM-aware codec
        io_string = io.StringIO(decoded_file)
        reader = csv.DictReader(io_string)
        cleaned_rows = []

        for row in reader:
            cleaned_row = {key.strip(): (value.strip() if isinstance(value, str) else value)
                           for key, value in row.items()}
            cleaned_rows.append(cleaned_row)

        return cleaned_rows

    def read_excel(self, file):
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        header = [cell.value for cell in ws[1]]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_data = dict(zip(header, row))
            rows.append(row_data)
        return rows

    def process_rows(self, rows):
        last_product = None
        for row_data in rows:
            name = row_data.get('Product Name')
            if name:
                category, subcategory, subsubcategory = get_or_create_category_hierarchy(
                    row_data)
                # Check if product with same name and subcategory exists
                if Product.objects.filter(name=name, subcategory=subcategory).exists():
                    last_product = None
                    continue
                product = create_product_from_row(
                    row_data, category, subcategory, subsubcategory)
                last_product = product
            product = last_product
            if not product:
                continue
            create_product_image(product, row_data)
        return Response({'message': 'Products imported successfully'}, status=201)


@method_decorator(csrf_exempt, name='dispatch')
class ProductGoogleSheetImportAPIView(APIView):
    def post(self, request):
        rows = request.data.get('rows')
        if not rows:
            return Response({'error': 'No data received'}, status=400)

        products = []
        current_product_data = None

        for row in rows:
            name = row.get('Product Name')
            if name:
                if current_product_data:
                    products.append(current_product_data)
                current_product_data = {
                    'product_data': row,
                    'images': [{
                        'Color': row.get('Color'),
                        'Stock': row.get('Stock (Color)', 0),
                        'Images': row.get('Images'),
                        'Image Alt Description': row.get('Image Alt Description')
                    }]
                }
            else:
                if current_product_data:
                    current_product_data['images'].append({
                        'Color': row.get('Color'),
                        'Stock': row.get('Stock (Color)', 0),
                        'Images': row.get('Images'),
                        'Image Alt Description': row.get('Image Alt Description')
                    })
        if current_product_data:
            products.append(current_product_data)

        for p in products:
            row = p['product_data']
            category, subcategory, subsubcategory = get_or_create_category_hierarchy(
                row)
            # Check if product with same name and subcategory exists
            if Product.objects.filter(name=row.get('Product Name'), subcategory=subcategory).exists():
                continue
            product = create_product_from_row(
                row, category, subcategory, subsubcategory)
            for img in p['images']:
                create_product_image(product, img)
        return Response({'message': '✅ Products imported successfully!'})
